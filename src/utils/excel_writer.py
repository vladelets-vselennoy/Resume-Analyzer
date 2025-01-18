import pandas as pd
from typing import List, Dict
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell  # Import MergedCell for the fix

class ExcelWriter:
    def __init__(self):
        # Define column groups
        self.column_groups = {
            'Contact Details': ['Name', 'Phone', 'Email', 'City', 'State'],
            'Education': ['University','Year_of_Study', 'Course', 'Discipline',  'CGPA'],
            'Skills & Experience': ['Key_Skills'],
            'AI Expertise': ['GenAI_Experience_Score', 'AI_ML_Experience_Score'],
            'Supporting Information': ['Internships','Projects', 'Certifications', 'Other_Achievements','Total_Experience','Matched_Jobs']
            
        }
        
        # Define metadata
        self.metadata = {
            'Project': 'Resume Analyzer',
            'Version': '1.0',
            'Generated': pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
        }

    def write_results(self, results: List[Dict], output_path: str) -> None:
        try:
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Convert results to DataFrame
                df = pd.DataFrame(results)
                
                # Write DataFrame starting from row 4 (after metadata)
                df.to_excel(writer, sheet_name='Results', startrow=3, index=False)
                
                # Get workbook and sheet
                workbook = writer.book
                worksheet = writer.sheets['Results']
                
                # Add metadata rows
                self._add_metadata(worksheet)
                
                # Add grouped headers
                self._add_group_headers(worksheet)
                
                # Apply styling
                self._apply_styling(worksheet, len(results))
                
                # Adjust column widths
                self._adjust_column_widths(worksheet)
                
            print(f"Results written successfully to {output_path}")
            self._open_file(output_path)
            
        except Exception as e:
            print(f"Error writing to Excel: {e}")
            raise
    def _open_file(self, file_path: str) -> None:
        """Open the file using the default system application."""
        try:
            if os.name == 'nt':  # Windows
                os.startfile(file_path)
            elif os.name == 'posix':  # macOS and Linux
                subprocess.call(['open', file_path])  # macOS
                # subprocess.call(['xdg-open', file_path])  # Linux
            else:
                print(f"Cannot automatically open file. Please open {file_path} manually.")
        except Exception as e:
            print(f"Error opening file: {e}")

    def _add_metadata(self, worksheet):
        """Add metadata rows at the top"""
        for idx, (key, value) in enumerate(self.metadata.items()):
            cell = worksheet.cell(row=idx+1, column=1)
            cell.value = f"{key}: {value}"
            cell.font = Font(bold=True)
            worksheet.merge_cells(start_row=idx+1, start_column=1, end_row=idx+1, end_column=5)

    def _add_group_headers(self, worksheet):
        """Add grouped headers"""
        current_col = 1
        row = 3
        
        for group, columns in self.column_groups.items():
            # Add group header
            start_col = current_col
            end_col = current_col + len(columns) - 1
            
            cell = worksheet.cell(row=row, column=start_col)
            cell.value = group
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
            worksheet.merge_cells(
                start_row=row, 
                start_column=start_col, 
                end_row=row, 
                end_column=end_col
            )
            
            # Add subheaders
            for idx, col in enumerate(columns, start=start_col):
                cell = worksheet.cell(row=row+1, column=idx)
                cell.value = col
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            current_col = end_col + 1

    def _apply_styling(self, worksheet, data_len):
        """Apply styling to the worksheet"""
        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders and alignment to data cells
        for row in range(4, data_len + 5):  # Data starts from row 5
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

    def _adjust_column_widths(self, worksheet):
        """Adjust column widths based on content"""
        for col in worksheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Use get_column_letter instead of column_letter
            
            for cell in col:
                # Skip merged cells entirely
                if isinstance(cell, MergedCell):
                    continue
                
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = min(adjusted_width, 50)
